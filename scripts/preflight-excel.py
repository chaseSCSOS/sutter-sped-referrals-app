import msoffcrypto
import io
import json
from openpyxl import load_workbook
from datetime import datetime

encrypted = open('Referral Log 2025 2026 2.xlsx', 'rb')
decrypted = io.BytesIO()
f = msoffcrypto.OfficeFile(encrypted)
f.load_key(password='2913')
f.decrypt(decrypted)
encrypted.close()
decrypted.seek(0)
wb = load_workbook(decrypted, data_only=True)

issues = []
stats = {}

VALID_SILOS = {'ASD', 'SD', 'NC', 'DHH', 'MD', 'OT'}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    non_empty = [r for r in rows if any(v is not None for v in r)]
    
    stats[sheet_name] = {
        'total_rows_including_blank': len(rows),
        'non_empty_rows': len(non_empty),
        'headers': headers
    }

    for i, row in enumerate(non_empty, 2):
        row_dict = dict(zip(headers, row))
        last = row_dict.get(' Last Name') or row_dict.get('Student Last Name')
        first = row_dict.get(' First Name') or row_dict.get('Student First Name')
        dob = row_dict.get('DOB') or row_dict.get('Students DOB')
        ref_date = row_dict.get('Referral Date') or row_dict.get('Date of Referral')

        if not last and not first:
            issues.append(f'{sheet_name} row {i}: SKIP - no name')
            continue

        if not dob:
            issues.append(f'{sheet_name} row {i}: WARN - missing DOB for {last}, {first}')
        if not ref_date:
            issues.append(f'{sheet_name} row {i}: WARN - missing referral date for {last}, {first}')

    # Silo value audit (SDC, BX)
    if sheet_name in ('SDC', 'BX'):
        silo_vals = {}
        for row in non_empty:
            row_dict = dict(zip(headers, row))
            s = row_dict.get('Silo')
            if s is not None:
                s_str = str(s).strip()
                silo_vals[s_str] = silo_vals.get(s_str, 0) + 1
                if s_str not in VALID_SILOS:
                    issues.append(f'{sheet_name}: WARN - unknown silo value: "{s_str}"')
        stats[sheet_name]['silo_value_counts'] = silo_vals

    # Form type combos (SDC)
    if sheet_name == 'SDC':
        form_combos = {}
        for row in non_empty:
            row_dict = dict(zip(headers, row))
            interim = bool(row_dict.get('Interim'))
            level2 = bool(row_dict.get('Level 2 Ref'))
            pip = bool(row_dict.get('PIP'))
            key = f'interim={interim}, level2={level2}, pip={pip}'
            form_combos[key] = form_combos.get(key, 0) + 1
        stats[sheet_name]['form_type_combos'] = form_combos

    # SEIS/Aeries value audit
    if sheet_name in ('SCIP', 'VIP'):
        seis_vals = {}
        aeries_vals = {}
        for row in non_empty:
            row_dict = dict(zip(headers, row))
            s = row_dict.get('In SEIS')
            a = row_dict.get('In Aeries')
            sv = str(s).strip() if s is not None else 'None'
            av = str(a).strip() if a is not None else 'None'
            seis_vals[sv] = seis_vals.get(sv, 0) + 1
            if a is not None:
                aeries_vals[av] = aeries_vals.get(av, 0) + 1
        stats[sheet_name]['seis_value_counts'] = seis_vals
        if aeries_vals:
            stats[sheet_name]['aeries_value_counts'] = aeries_vals

    # CUM date format audit (SDC, BX)
    if sheet_name in ('SDC', 'BX'):
        cum_samples = []
        for row in non_empty:
            row_dict = dict(zip(headers, row))
            cum_recv = row_dict.get('Date CUM Received and Sent to')
            if cum_recv and isinstance(cum_recv, str):
                cum_samples.append(cum_recv)
        if cum_samples:
            stats[sheet_name]['cum_string_samples'] = cum_samples[:10]

print('=== EXCEL PRE-FLIGHT RESULTS ===')
print(json.dumps(stats, indent=2, default=str))
print('\n=== ISSUES ===')
for issue in issues:
    print(' ', issue)
print(f'\nTotal issues: {len(issues)}')
