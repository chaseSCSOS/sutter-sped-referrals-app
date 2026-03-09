import msoffcrypto
import io
from openpyxl import load_workbook

encrypted = open('Referral Log 2025 2026 2.xlsx', 'rb')
decrypted = io.BytesIO()
f = msoffcrypto.OfficeFile(encrypted)
f.load_key(password='2913')
f.decrypt(decrypted)
encrypted.close()
decrypted.seek(0)
wb = load_workbook(decrypted, data_only=True)

# 1. SDC trailing rows (80-92) - understand incomplete records
ws = wb['SDC']
headers = [cell.value for cell in ws[1]]
print('=== SDC rows 80-92 (end of data) ===')
for i, row in enumerate(ws.iter_rows(min_row=80, max_row=92, values_only=True), 80):
    row_dict = dict(zip(headers, row))
    non_null = {k: v for k, v in row_dict.items() if v is not None and k is not None}
    print(f'  Row {i}: {non_null}')

# 2. BX rows 1-10 - understand the skip rows
ws2 = wb['BX']
headers2 = [cell.value for cell in ws2[1]]
print('\n=== BX rows 1-10 ===')
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    print(f'  Row {i}: {list(row)}')

# 3. SDC rows with no form type checkbox
ws3 = wb['SDC']
headers3 = [cell.value for cell in ws3[1]]
print('\n=== SDC rows with NO form type checked (no Interim, Level2, or PIP) ===')
for i, row in enumerate(ws3.iter_rows(min_row=2, values_only=True), 2):
    row_dict = dict(zip(headers3, row))
    if not row_dict.get('Interim') and not row_dict.get('Level 2 Ref') and not row_dict.get('PIP'):
        if any(v is not None for v in row):
            last = row_dict.get('Student Last Name')
            first = row_dict.get('Student First Name')
            silo = row_dict.get('Silo')
            dob = row_dict.get('Students DOB')
            print(f'  Row {i}: {last}, {first} | silo={silo} | dob={dob}')

# 4. Check CUM Requested column - what kind of values appear there
print('\n=== SDC CUM Requested column sample values ===')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=20, values_only=True), 2):
    row_dict = dict(zip(headers, row))
    cum_req = row_dict.get('CUM Requested ')
    last = row_dict.get('Student Last Name')
    if last:
        print(f'  Row {i}: last={last}, CUM Requested={repr(cum_req)}')
