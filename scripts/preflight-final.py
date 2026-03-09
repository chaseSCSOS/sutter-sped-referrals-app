import msoffcrypto
import io
from openpyxl import load_workbook
from datetime import datetime

encrypted = open('Referral Log 2025 2026 2.xlsx', 'rb')
decrypted = io.BytesIO()
f = msoffcrypto.OfficeFile(encrypted)
f.load_key(password='2913')
f.decrypt(decrypted)
encrypted.close()
decrypted.seek(0)
wb = load_workbook(decrypted, data_only=True)

def is_valid_row(row_dict, last_key, first_key, dob_key):
    last = row_dict.get(last_key)
    first = row_dict.get(first_key)
    dob = row_dict.get(dob_key)
    # Must have at least last name and a real DOB (datetime, not a number or short string)
    if not last:
        return False
    if not isinstance(dob, datetime):
        return False
    return True

print('=== FINAL IMPORTABLE ROW COUNTS ===\n')

# SDC
ws = wb['SDC']
headers = [cell.value for cell in ws[1]]
valid = []
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    row_dict = dict(zip(headers, row))
    if is_valid_row(row_dict, 'Student Last Name', 'Student First Name', 'Students DOB'):
        valid.append((i, row_dict))
print(f'SDC: {len(valid)} importable rows')
# Show form type breakdown
ft = {'INTERIM':0,'LEVEL_II':0,'PIP_ONLY':0,'NONE':0}
for _, r in valid:
    if r.get('Interim'): ft['INTERIM'] += 1
    elif r.get('Level 2 Ref'): ft['LEVEL_II'] += 1
    elif r.get('PIP'): ft['PIP_ONLY'] += 1
    else: ft['NONE'] += 1
print(f'  Form types: {ft}')

# BX
ws2 = wb['BX']
headers2 = [cell.value for cell in ws2[1]]
valid2 = []
for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), 2):
    row_dict = dict(zip(headers2, row))
    if is_valid_row(row_dict, 'Student Last Name', 'Student First Name', 'Students DOB'):
        valid2.append((i, row_dict))
print(f'BX: {len(valid2)} importable rows')
for i, r in valid2:
    print(f'  Row {i}: {r.get("Student Last Name")}, {r.get("Student First Name")}')

# SCIP
ws3 = wb['SCIP']
headers3 = [cell.value for cell in ws3[1]]
valid3 = []
for i, row in enumerate(ws3.iter_rows(min_row=2, values_only=True), 2):
    row_dict = dict(zip(headers3, row))
    if is_valid_row(row_dict, ' Last Name', ' First Name', 'DOB'):
        valid3.append((i, row_dict))
print(f'SCIP: {len(valid3)} importable rows')

# VIP
ws4 = wb['VIP']
headers4 = [cell.value for cell in ws4[1]]
valid4 = []
for i, row in enumerate(ws4.iter_rows(min_row=2, values_only=True), 2):
    row_dict = dict(zip(headers4, row))
    if is_valid_row(row_dict, ' Last Name', ' First Name', 'DOB'):
        valid4.append((i, row_dict))
print(f'VIP: {len(valid4)} importable rows')

# DHH
ws5 = wb['DHH']
headers5 = [cell.value for cell in ws5[1]]
valid5 = []
for i, row in enumerate(ws5.iter_rows(min_row=2, values_only=True), 2):
    row_dict = dict(zip(headers5, row))
    if is_valid_row(row_dict, ' Last Name', ' First Name', 'DOB'):
        valid5.append((i, row_dict))
print(f'DHH: {len(valid5)} importable rows')

total = len(valid) + len(valid2) + len(valid3) + len(valid4) + len(valid5)
print(f'\nTOTAL IMPORTABLE: {total}')

# CUM Requested column: show all unique values to understand what it stores
print('\n=== SDC "CUM Requested" column — all unique values ===')
cum_vals = {}
for _, r in valid:
    v = r.get('CUM Requested ')
    if v is not None:
        k = str(v) if not isinstance(v, datetime) else 'datetime:' + str(v.date())
        cum_vals[k] = cum_vals.get(k, 0) + 1
for k, c in sorted(cum_vals.items(), key=lambda x: -x[1]):
    print(f'  {repr(k)}: {c}')

# Show a few complete BX records for verification
print('\n=== BX complete records ===')
for i, r in valid2:
    print(f'  Row {i}: {r}')

# Check SDC row 61 (Bacchas, Averee) - no form type
print('\n=== SDC row with no form type (Bacchas) ===')
for i, r in valid:
    if r.get('Student Last Name') == 'Bacchas':
        print(f'  Row {i}: {r}')
